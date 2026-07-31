"""
test_e2e.py - end-to-end test: FRESH summaries -> judge -> compare vs Abe.

Input xlsx = fresh Claude-generated summaries from live Close pulls
(same 5 columns as the audit sheets). For each row:
  embed -> retrieve top matches EXCLUDING the lead itself -> judge ->
  print AI verdict next to Abe's stored verdict from Supabase.

Usage:
    python3 test_e2e.py makar_e2e.xlsx

NOTE: leads move. If the lead's state changed since Abe judged it
(new inbound, status moved), an AI/Abe difference may mean the AI is
RIGHT about the new state. The output shows both so you can eyeball.
"""
import sys, json, re, time
import openpyxl
from voyageai.error import RateLimitError
from common import sb, claude, embed, CLAUDE_MODEL


def parse_facts(activity):
    """Extract the FACTS line into a dict, or {} if absent."""
    m = re.search(r"FACTS: ([^\n]+)", str(activity or ""))
    if not m:
        return {}
    f = {}
    for part in m.group(1).split(" | "):
        if "=" in part:
            k, v = part.split("=", 1)
            f[k.strip()] = v.strip()
    return f


def apply_fact_overrides(d, activity):
    """Fact fields are computed, not judged. Override the model where FACTS exist."""
    f = parse_facts(activity)
    if not f:
        return d
    try:
        recent = int(f.get("max_call_sec_90d", f.get("owner_max_call_sec", 0)))
        d["convo_2min"] = "Yes" if recent >= 120 else "No"   # strict 2min rule
    except ValueError:
        pass
    if "rev_filled" in f:
        d["rev_filled"] = "Yes" if f["rev_filled"].startswith("Yes") else "No"
    return d

PROMPT = """You are Abe, founder of an MCA brokerage, auditing how a lead was handled.
Judge the LAST USER WHO COMMUNICATED with the merchant (FACTS judged_user) - not the Lead Owner field.

ABE'S OFFICIAL CRITERIA:
1. convo_2min: did the judged user have a call over 2 minutes (>=120s) within the window?
   Mechanical: auto-set in code from FACTS max_call_sec_90d.
2. rep_error: EMPTY ("") if there is no error - that is the common case. Assign only when one clearly applies:
   * "No connect to owner" - judged user never reached the decision-maker live.
   * "Status-too long" - something live is parked in a slow bucket for weeks (see days_in_current_status).
   * "Sales Rep Conversation Skill" - a convo happened and the rep clearly botched it (ignored stated channel_pref, broke a merchant-agreed timeline, never asked for app/time on a live ask).
   * "Didnt collect information and was possible" - real convo happened, key info was gettable, none captured.
   When torn, choose "".
3. task_verdict: A task SHOULD exist only when the merchant specified an exact time/date to talk.
   Use FACTS task_by_last_user (was a task put in by the judged user):
   * merchant gave a specific time/date + task_by_last_user=Yes -> "Yes - Necessary"
   * merchant gave a specific time/date + task_by_last_user=No  -> "No - Should Have"
   * no specific time/date given + task_by_last_user=Yes -> "Yes - Unnecessary"
   * no specific time/date given + task_by_last_user=No  -> "No - Not Needed"
4. meaningful_touch: Yes when there was a 2min+ call (or equivalent substantive exchange) that covered what the company does and/or its revenue AND what Getty can offer (product/terms discussed). A CALL TRANSCRIPT showing revenue, docs requested, or product terms discussed = Yes. Dials, voicemails, one-word replies = No.
5. next_steps: from the conversation + email content. Positive statuses = 3 Day / 14 Day / 40 Day / High Priority. Pick EXACTLY ONE unless two are truly required:
   * live deal or honored merchant timeline -> "Follow up"
   * live ask, no app yet -> "Get App"
   * docs/statements promised or the natural next ask -> "Get Banks"
   * dead / DQ / opted-out / doesn't belong in a positive status -> "Remove from positive"
   * lead alive but bucket or follow-up date wrong -> "Adjust follow up date or status"
   * merchant was interested but the rep didn't ask questions / no follow-up email, SMS, or call to continue the conversation -> "Sales rep conversation training"
6. rev_filled: mechanical - copy FACTS rev_filled (the Rep Reported Revenue custom field, empty or "0. N/A" = No). Auto-set in code.
days_in_current_status=-1 means UNKNOWN - never treat it as evidence of anything.
Also: correct_status judges the STATUS bucket against what the MERCHANT said - look at call/SMS/email CONTENT and what the merchant replied about next steps, not just activity stats:
   * "busy now" / "call me later" / "call me in 1-3 days" -> belongs in 3 Day Follow up
   * "call me in a few weeks" (1-3 weeks) -> belongs in 14 Day Follow up
   * "call me in a month+" / "4-6 weeks" / "about a month" / any season or quarter reference ("summer", "fall", "winter", "next quarter") / "interested but not now" / anything implying a callback ~30+ days out -> belongs in 40 Day Follow up
   * live active deal being worked -> High Priority
   Status matches the merchant's stated timeline = Yes. Wrong horizon (e.g. a "few weeks" ask sitting in 40 Day, a live deal in 40 Day, a dead/opted-out lead in any positive bucket) = No. A dead lead correctly parked/demoted = Yes.

ABE'S CALIBRATION LANES (learned from his corrections - apply exactly):
- Merchant-set SLOW timeline (quarterly check-ins, "I'll text if things change", "I'll contact you"): correct_status = Yes AND next_steps = "Remove from positive" - honored slow timelines come OFF the 3/14/40 cadence entirely.
- Ghosted promise, FRESH (days to ~6 weeks, channels alive): "Follow up" (or "Get Banks" if docs outstanding). Ghosted promise STALE (2+ months, merchant dodging "can't talk now", or contact channels dead): "Remove from positive".
- Live deal where materials/app/product email was ALREADY SENT: next_steps = "Follow up", not "Get App". "Get App" is for a live ask where nothing has been sent yet or the app must be re-driven after a reversal.
- Sudden reversal on a live deal ("No thx" days after full buy-in, unexplained): status may STAY High Priority (objection to uncover, not a real no); next_steps = "Get App".
- Rep imperfections mentioned in passing (a broken cadence, a missed nuance) do NOT earn a rep_error when the lead was handled reasonably overall - rep_error is assigned only when the error IS the story of the lead. Abe leaves it blank far more often than not.
- Past funded client eligible for renewal sitting in a cold generic cadence: correct_status = No (belongs on a renewal track, not 40 Day).
- Default toward correct_status = Yes for any DEFENSIBLE bucket. Overturn only on a clear contradiction: a live signal in a slow bucket, a dead/opted-out lead in a positive bucket, or a wrong horizon vs the merchant's stated words.

CRITICAL EVIDENCE RULES:
- A merchant who agreed to next steps (statements, app, docs) and then went SILENT is an unfinished promise, NOT a dead lead. Silence after a promise -> "Follow up" (or "Get Banks" if docs are the outstanding item). NEVER "Remove from positive" for ghosting alone.
- "Remove from positive" requires an EXPLICIT dead signal: stated decline, STOP/opt-out, DQ math, or all contact channels confirmed dead. Never infer funded-elsewhere/closed from ambiguous one-liners ('ok', 'it done') - require an explicit statement.
- CALL TRANSCRIPT lines are the primary evidence - what was actually said beats duration stats. If a transcript shows business/revenue discussed + Getty product discussed, meaningful_touch=Yes. Merchant statements in transcripts (amounts, promises, timelines like 'this weekend'/'next week'/'in the fall') are the source of truth for correct_status and next_steps.
- A merchant-agreed future reconnect ("get back in touch in the fall", "check in after summer") VALIDATES the matching bucket (40 Day) as correct_status=Yes - even if a product-fit question (e.g. nonprofit) is unresolved. Honoring an agreed timeline is never a wrong status.
- High Priority is correct when a recent substantive call shows a live deal being worked; do not overturn HP or a follow-up bucket merely because days have passed with no reply.
- Do not mark correct_status No solely because a lead is old, quiet, or thin - Abe's default for a plausible bucket with no contradicting merchant statement is Yes.

Here is how you judged the most similar past situations (your validated reviews):
{examples}

New lead to judge:
Status: {status}
Owner activity (90d): {activity}
Owner meaningful touch: {touch}

Return ONLY JSON, no markdown fences:
{{"correct_status":"...","convo_2min":"...","rep_error":"...","task_verdict":"...",
 "meaningful_touch":"...","next_steps":"...","rev_filled":"...","note":"..."}}"""

FIELDS = [("correct_status", "abe_correct_status"),
          ("convo_2min", "abe_owner_convo_2min"),
          ("rep_error", "abe_rep_error_category"),
          ("task_verdict", "abe_task_verdict"),
          ("meaningful_touch", "abe_meaningful_touch"),
          ("next_steps", "abe_next_steps"),
          ("rev_filled", "abe_rev_filled")]


def embed_throttled(texts, kind):
    for _ in range(6):
        try:
            return embed(texts, kind=kind)
        except RateLimitError:
            print("  ...rate limited, waiting 21s")
            time.sleep(21)
    raise SystemExit("Voyage rate limit persisted.")


def main(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "") for c in ws[1]]

    def col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k.lower() in h.lower():
                    return i
        return None
    i_name, i_id = col("Lead") or 0, col("Lead ID")
    i_status, i_act = col("Status"), col("Owner Activity")
    i_touch = col("Meaningful touch", "Owner Meaningful")
    scores, n = {}, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or str(row[0]) == "ERROR":
            continue
        name, lead_id = row[i_name], row[i_id]
        status = row[i_status] if i_status is not None else ""
        activity = row[i_act] if i_act is not None else ""
        touch = row[i_touch] if i_touch is not None else ""

        situation = (f"Status: {status}\nOwner activity (90d): {activity}\n"
                     f"Owner meaningful touch: {touch}")
        q = embed_throttled([situation], kind="query")[0]
        res = sb.rpc("match_abe_notes", {"query_embedding": q, "match_count": 8}).execute()
        matches = [m for m in res.data if m["lead_name"].strip().lower()
                   not in str(name).strip().lower()
                   and str(name).strip().lower() not in m["lead_name"].strip().lower()][:6]
        top_sim = matches[0]["similarity"] if matches else 0

        examples = "\n".join(
            f"- {m['status_before']} | {(m['owner_activity'] or '')[:200]}\n"
            f"  YOUR VERDICT -> Correct status: {m['abe_correct_status']} | "
            f"Rep error: {m['abe_rep_error_category'] or 'none'} | "
            f"Task: {m['abe_task_verdict']} | Meaningful: {m['abe_meaningful_touch']} | "
            f"Next: {m['abe_next_steps']}"
            for m in matches)

        msg = claude.messages.create(
            model=CLAUDE_MODEL, max_tokens=450,
            messages=[{"role": "user", "content": PROMPT.format(
                examples=examples, status=status, activity=activity, touch=touch)}])
        txt = msg.content[0].text.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
        try:
            d = json.loads(txt)
        except json.JSONDecodeError:
            print(f"!! {name}: bad JSON\n{txt}\n")
            continue
        d = apply_fact_overrides(d, activity)

        abe = sb.table("abe_notes").select(
            ", ".join(c for _, c in FIELDS)).eq("lead_id", lead_id).execute().data
        abe = abe[0] if abe else {}

        print(f"\n{'='*60}\n{name}  (top sim {top_sim:.2f})")
        print(f"{'field':<18}{'AI (fresh state)':<26}{'ABE (at audit time)'}")
        for k, col in FIELDS:
            ai = str(d.get(k) or "")
            av = str(abe.get(col) or "") if abe else "(not in cabinet)"
            mark = " " if ai.strip() == av.strip() else "*"
            print(f"{mark}{k:<17}{ai:<26}{av}")
        print(f"AI note: {d.get('note','')}")
        if not abe:
            print("(no stored Abe verdict for this lead_id)")
        else:
            n += 1
            for k, c in FIELDS:
                scores[k] = scores.get(k, 0) + (1 if str(d.get(k) or "").strip() == str(abe.get(c) or "").strip() else 0)
    if n:
        print(f"\n{'='*60}\nSCORECARD ({n} leads with stored verdicts)")
        for k, _ in FIELDS:
            print(f"  {k:<18} {100*scores.get(k,0)/n:5.1f}%")
        overall = 100 * sum(scores.values()) / (len(FIELDS) * n)
        print(f"  {'OVERALL':<18} {overall:5.1f}%")
    print("\n* = differs. Remember: if the lead moved since Abe judged it,")
    print("the AI may be correctly judging the NEW state - eyeball each *.")


if __name__ == "__main__":
    main(sys.argv[1])
