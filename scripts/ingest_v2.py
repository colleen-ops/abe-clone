"""
ingest_v2.py — load the Makar2-format sheet (Abe's VALIDATED audit) into Supabase.

Usage:
    python ingest_v2.py data/makar2.xlsx
    (export the Makar2 tab: in Google Sheets open that tab -> File -> Download -> .xlsx)

Makar2 columns expected (header row):
    Lead | Lead ID | Status | Owner Activity Summary(90d) | Owner Meaningful Touch? |
    Overall Activity | Overall Meaningful touch? | Abe Review (Y/N) | Abe note |
    Correct Status | Lead Owner Convo 2min> | Rep Error |
    Was a Task put in? and if so, was it necessary? | Meaningful Touch? |
    Next Steps | Rep Reported Revenue Filled Out | <free-text notes col>

Only rows where Abe validated (Correct Status filled) are ingested.
Embeds the SITUATION = Status + Owner Activity Summary + Owner Meaningful Touch.
Safe to re-run: skips lead_ids already in the DB.
"""
import sys
import openpyxl
from common import sb, embed


def find_col(headers, *keys):
    """First header containing any key (case-insensitive)."""
    for key in keys:
        for h in headers:
            if h and key.lower() in str(h).lower():
                return h
    return None


def main(path):
    ws = openpyxl.load_workbook(path).active
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    col = {
        "owner_activity": find_col(headers, "Owner Activity"),
        "owner_touch": find_col(headers, "Owner Meaningful"),
        "overall_activity": find_col(headers, "Overall Activity"),
        "overall_touch": find_col(headers, "Overall Meaningful"),
        "abe_note": find_col(headers, "Abe note"),
        "correct_status": find_col(headers, "Correct Status"),
        "convo_2min": find_col(headers, "Convo 2min", "Lead Owner Convo"),
        "rep_error": find_col(headers, "Rep Error"),
        "task": find_col(headers, "Task put in"),
        "abe_touch": find_col(headers, "Meaningful Touch?"),
        "next_steps": find_col(headers, "Next Steps"),
        "rev_filled": find_col(headers, "Revenue Filled"),
    }
    # free-text = last non-empty header after Rev Filled, if any
    freetext_col = headers[-1] if headers[-1] and headers[-1] != col["rev_filled"] else None

    existing = {r["lead_id"] for r in sb.table("abe_notes").select("lead_id").execute().data}
    g = lambda r, k: (str(r.get(col[k]) or "").strip()) if col[k] else ""

    done = skipped = 0
    for r in rows:
        lead_id = str(r.get("Lead ID") or "").strip()
        verdict = g(r, "correct_status")
        if not lead_id.startswith("lead_") or not verdict:
            skipped += 1
            continue
        if lead_id in existing:
            skipped += 1
            continue

        situation = (
            f"Status: {r.get('Status','')}\n"
            f"Owner activity (90d): {g(r,'owner_activity')}\n"
            f"Owner meaningful touch: {g(r,'owner_touch')}"
        )
        sb.table("abe_notes").insert({
            "lead_id": lead_id,
            "lead_name": r.get("Lead"),
            "status_before": r.get("Status"),
            "owner_activity": g(r, "owner_activity"),
            "owner_meaningful_touch": g(r, "owner_touch"),
            "overall_activity": g(r, "overall_activity"),
            "overall_meaningful_touch": g(r, "overall_touch"),
            "abe_correct_status": verdict,
            "abe_owner_convo_2min": g(r, "convo_2min"),
            "abe_rep_error_category": g(r, "rep_error"),
            "abe_task_verdict": g(r, "task"),
            "abe_meaningful_touch": g(r, "abe_touch"),
            "abe_next_steps": g(r, "next_steps"),
            "abe_rev_filled": g(r, "rev_filled"),
            "abe_freetext": (str(r.get(freetext_col) or "").strip() if freetext_col else "") or g(r, "abe_note"),
            "full_note": " | ".join(filter(None, [
                f"Correct status: {verdict}", g(r, "rep_error"),
                f"Next: {g(r,'next_steps')}", g(r, "abe_note")])),
            "embedding": embed([situation], kind="document")[0],
        }).execute()
        done += 1
        print(f"  ✓ {r.get('Lead')} — Correct status: {verdict} | Next: {g(r,'next_steps')}")

    print(f"\nDone. {done} validated rows ingested, {skipped} skipped (no Abe verdict / already in DB / blank).")


if __name__ == "__main__":
    main(sys.argv[1])
