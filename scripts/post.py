"""
post.py — after Abe reviews a *_DRAFTS.xlsx:
  1. posts approved notes to the lead in Close (as a Note activity)
  2. ingests them back into the vector DB (flywheel)

Usage:
    python post.py data/next_batch_DRAFTS.xlsx            # dry run (prints, posts nothing)
    python post.py data/next_batch_DRAFTS.xlsx --send     # actually posts to Close

Rules:
  Abe Approve = Y            -> use AI Note as-is
  Abe Approve = edit         -> use Abe Final Note
  blank / N                  -> skipped
"""
import sys, requests
from common import sb, embed, lead_context, CLOSE_KEY

CLOSE_NOTE_URL = "https://api.close.com/api/v1/activity/note/"


def post_to_close(lead_id, note):
    r = requests.post(
        CLOSE_NOTE_URL,
        auth=(CLOSE_KEY, ""),
        json={"lead_id": lead_id, "note": f"[ABE AUDIT]\n{note}"},
        timeout=30,
    )
    r.raise_for_status()


def parse_part(note, tag):
    for line in note.splitlines():
        if line.startswith(tag + ":"):
            return line.split(":", 1)[1].strip()
    return ""


def main(path, send=False):
    import openpyxl
    ws = openpyxl.load_workbook(path).active
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    existing = {r["lead_id"] for r in sb.table("abe_notes").select("lead_id").execute().data}
    posted = skipped = 0

    for r in rows:
        approve = str(r.get("Abe Approve (Y/N/edit)") or "").strip().lower()
        if approve in ("", "n"):
            skipped += 1
            continue
        note = (r.get("Abe Final Note") or "").strip() if approve == "edit" else (r.get("AI Note") or "").strip()
        if not note or not r.get("Lead ID"):
            skipped += 1
            continue

        if send:
            post_to_close(r["Lead ID"], note)

        if r["Lead ID"] not in existing:
            ctx = lead_context(r)
            sb.table("abe_notes").insert({
                "lead_id": r["Lead ID"],
                "lead_name": r.get("Lead"),
                "status_before": r.get("Status"),
                "activity": r.get("Activity"),
                "meaningful_touch": r.get("Meaningful touch?"),
                "verdict": parse_part(note, "VERDICT"),
                "evidence": parse_part(note, "EVIDENCE"),
                "rep_error": parse_part(note, "REP ERROR"),
                "play": parse_part(note, "PLAY"),
                "rule": parse_part(note, "RULE"),
                "full_note": note,
                "embedding": embed([ctx], kind="document")[0],
            }).execute()

        posted += 1
        print(f"  ✓ {r.get('Lead')} {'(POSTED)' if send else '(dry run)'}")

    mode = "SENT to Close" if send else "DRY RUN — re-run with --send to post"
    print(f"\n{posted} notes processed, {skipped} skipped. {mode}")


if __name__ == "__main__":
    main(sys.argv[1], send="--send" in sys.argv)
